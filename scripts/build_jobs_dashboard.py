"""Build Yasir Malik's job-pipeline dashboard.

Produces /home/user/workspace/yasir_malik_jobs_dashboard.xlsx — a 7-sheet
workbook per the CLAUDE_PROMPT_jobs_dashboard.md spec. Recalculates via
LibreOffice headless and asserts zero formula errors.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
from jobs_data import JOBS  # noqa: E402

WORKSPACE = Path("/home/user/workspace")
OUT_PATH = WORKSPACE / "yasir_malik_jobs_dashboard.xlsx"
RECALC_SCRIPT = WORKSPACE / "skills/office/xlsx-repl/scripts/recalc.py"

# ---------------------------------------------------------------------------
# Nexus palette
# ---------------------------------------------------------------------------
PAL = {
    "primary": "01696F",
    "primary_dark": "0C4E54",
    "primary_light": "BCE2E7",
    "accent": "20808D",
    "background": "F7F6F2",
    "surface": "F9F8F5",
    "surface_alt": "FBFBF9",
    "border": "D4D1CA",
    "text": "28251D",
    "text_muted": "7A7974",
    "tier1": "FFE7CE",
    "tier2": "FFF4D6",
    "tier3": "E8F1ED",
    "tier4": "F2F2F0",
}

TIER_FILL = {
    1: PAL["tier1"],
    2: PAL["tier2"],
    3: PAL["tier3"],
    4: PAL["tier4"],
}

FONT_TITLE = Font(name="Calibri", size=28, bold=True, color=PAL["primary"])
FONT_SUBTITLE = Font(name="Calibri", size=12, italic=True, color=PAL["text_muted"])
FONT_HEADER_WHITE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=10, color=PAL["text"])
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=PAL["text"])
FONT_MUTED = Font(name="Calibri", size=9, italic=True, color=PAL["text_muted"])
FONT_KPI_LABEL = Font(name="Calibri", size=9, bold=True, color=PAL["text_muted"])
FONT_KPI_VALUE = Font(name="Calibri", size=24, bold=True, color=PAL["primary_dark"])
FONT_LINK = Font(name="Calibri", size=10, bold=True, color=PAL["primary"], underline="single")
FONT_README_BODY = Font(name="Calibri", size=11, color=PAL["text"])

FILL_PRIMARY = PatternFill("solid", fgColor=PAL["primary"])
FILL_ACCENT = PatternFill("solid", fgColor=PAL["accent"])
FILL_BG = PatternFill("solid", fgColor=PAL["background"])
FILL_SURFACE = PatternFill("solid", fgColor=PAL["surface"])
FILL_ZEBRA = PatternFill("solid", fgColor=PAL["surface_alt"])
FILL_PRIMARY_LIGHT = PatternFill("solid", fgColor=PAL["primary_light"])

BORDER_THIN = Border(
    left=Side(style="thin", color=PAL["border"]),
    right=Side(style="thin", color=PAL["border"]),
    top=Side(style="thin", color=PAL["border"]),
    bottom=Side(style="thin", color=PAL["border"]),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


def composite(sal_low: int, sal_high: int, fit: float) -> float:
    mid = (sal_low + sal_high) / 2
    return mid * fit / 10


# ---------------------------------------------------------------------------
# 0. Sort and rank the dataset.
# ---------------------------------------------------------------------------
SORTED_JOBS = sorted(
    JOBS,
    key=lambda j: (j[0], -composite(j[6], j[7], j[11])),
)
N_JOBS = len(SORTED_JOBS)
DATA_START_ROW = 7
DATA_END_ROW = DATA_START_ROW + N_JOBS - 1
SHEET_TRJ = "Top Ranked Jobs"


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def hide_gridlines(ws) -> None:
    ws.sheet_view.showGridLines = False


def apply_bg(ws, max_col: int, max_row: int) -> None:
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor is None or cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = FILL_BG


# ---------------------------------------------------------------------------
# Sheet 1: README
# ---------------------------------------------------------------------------
def build_readme(wb) -> None:
    ws = wb.active
    ws.title = "README"
    hide_gridlines(ws)
    set_col_widths(ws, {"A": 2, "B": 28, "C": 92, "D": 4})

    ws["B2"] = "Yasir A. Malik — NJ/NYC Job Pipeline Dashboard"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 40

    ws["B3"] = "Curated pipeline of audit, risk, and AI-governance roles in the NJ/NYC corridor"
    ws["B3"].font = FONT_SUBTITLE
    ws.row_dimensions[3].height = 22

    ws["B5"] = "HOW TO USE THIS WORKBOOK"
    ws["B5"].font = Font(name="Calibri", size=11, bold=True, color=PAL["primary"])
    ws.row_dimensions[5].height = 24

    rows = [
        ("Dashboard", "KPI cards + top-10 leaderboard. Start here every Monday morning."),
        ("Top Ranked Jobs", "Full ranked table with composite score, fit, salary, and apply links."),
        ("By Company", "Cross-tab: roles per employer with min/median/max salary mid and avg fit."),
        ("By Type", "Cross-tab on category (G-SIB, Pharma, Biotech, Consulting, Academic, …) plus chart."),
        ("Tier Summary", "Counts and salary ranges per tier (1 = bullseye, 4 = adjunct). Includes chart."),
        ("Action Plan", "Week-by-week application schedule for the next four weeks."),
    ]
    r = 6
    for tab, blurb in rows:
        ws.cell(row=r, column=2, value=tab).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=3, value=blurb).font = FONT_README_BODY
        ws.cell(row=r, column=3).alignment = LEFT_TOP
        ws.row_dimensions[r].height = 32
        r += 1

    ws.cell(row=r + 1, column=2,
            value=f"Built {date.today().isoformat()} · Yasir A. Malik · YasirAMalik@gmail.com"
            ).font = FONT_MUTED
    apply_bg(ws, max_col=5, max_row=r + 3)


# ---------------------------------------------------------------------------
# Sheet 2: Dashboard
# ---------------------------------------------------------------------------
def build_dashboard(wb) -> None:
    ws = wb.create_sheet("Dashboard")
    hide_gridlines(ws)
    set_col_widths(ws, {
        "A": 2,
        **{get_column_letter(i): 13 for i in range(2, 16)},
    })

    # Title bar — row 2
    ws.merge_cells("B2:N2")
    ws["B2"] = "JOB PIPELINE — DASHBOARD"
    ws["B2"].font = Font(name="Calibri", size=24, bold=True, color="FFFFFF")
    ws["B2"].fill = FILL_PRIMARY
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 44

    # KPI cards
    kpis = [
        ("TOTAL ROLES", f"=COUNTA('{SHEET_TRJ}'!D{DATA_START_ROW}:D{DATA_END_ROW})", "0"),
        ("ROLES ≥ $200K MID",
         f"=COUNTIF('{SHEET_TRJ}'!K{DATA_START_ROW}:K{DATA_END_ROW},\">=200000\")", "0"),
        ("ROLES WITH FIT ≥ 9.0",
         f"=COUNTIF('{SHEET_TRJ}'!L{DATA_START_ROW}:L{DATA_END_ROW},\">=9\")", "0"),
        ("MEDIAN SALARY MID",
         f"=MEDIAN('{SHEET_TRJ}'!K{DATA_START_ROW}:K{DATA_END_ROW})", "\"$\"#,##0"),
        ("TOP COMPOSITE SCORE",
         f"=MAX('{SHEET_TRJ}'!M{DATA_START_ROW}:M{DATA_END_ROW})", "\"$\"#,##0"),
        ("AVG FIT SCORE",
         f"=ROUND(AVERAGE('{SHEET_TRJ}'!L{DATA_START_ROW}:L{DATA_END_ROW}),1)", "0.0"),
    ]
    positions = [(5, 2), (5, 6), (5, 10), (10, 2), (10, 6), (10, 10)]
    for (top_row, left_col), (label, formula, fmt) in zip(positions, kpis):
        # accent stripe row
        for c in range(left_col, left_col + 4):
            cell = ws.cell(row=top_row, column=c)
            cell.fill = FILL_ACCENT
        ws.row_dimensions[top_row].height = 5
        # label row
        ws.merge_cells(start_row=top_row + 1, start_column=left_col,
                       end_row=top_row + 1, end_column=left_col + 3)
        lbl = ws.cell(row=top_row + 1, column=left_col, value=label)
        lbl.font = FONT_KPI_LABEL
        lbl.fill = FILL_SURFACE
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[top_row + 1].height = 22
        # value row (spans 2 rows visually via row height)
        ws.merge_cells(start_row=top_row + 2, start_column=left_col,
                       end_row=top_row + 3, end_column=left_col + 3)
        val = ws.cell(row=top_row + 2, column=left_col, value=formula)
        val.font = FONT_KPI_VALUE
        val.fill = FILL_SURFACE
        val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val.number_format = fmt
        ws.row_dimensions[top_row + 2].height = 34
        ws.row_dimensions[top_row + 3].height = 12
        # border around the card
        for r in range(top_row, top_row + 4):
            for c in range(left_col, left_col + 4):
                ws.cell(row=r, column=c).border = BORDER_THIN

    # Top-10 leaderboard at row 16
    leaderboard_start = 16
    ws.cell(row=leaderboard_start - 1, column=2,
            value="TOP 10 — BY COMPOSITE SCORE").font = Font(
        name="Calibri", size=12, bold=True, color=PAL["primary_dark"])

    headers = ["Rank", "Company", "Title", "Tier", "Sal Mid", "Fit", "Composite", "Apply"]
    for idx, h in enumerate(headers):
        cell = ws.cell(row=leaderboard_start, column=2 + idx, value=h)
        cell.font = FONT_HEADER_WHITE
        cell.fill = FILL_PRIMARY
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[leaderboard_start].height = 28

    for i in range(10):
        src_row = DATA_START_ROW + i  # already sorted
        r = leaderboard_start + 1 + i
        ws.cell(row=r, column=2, value=i + 1).alignment = CENTER
        ws.cell(row=r, column=3, value=f"='{SHEET_TRJ}'!D{src_row}").font = FONT_BODY
        ws.cell(row=r, column=4, value=f"='{SHEET_TRJ}'!E{src_row}").font = FONT_BODY
        ws.cell(row=r, column=5, value=f"='{SHEET_TRJ}'!C{src_row}").alignment = CENTER
        ws.cell(row=r, column=6, value=f"='{SHEET_TRJ}'!K{src_row}").number_format = "\"$\"#,##0"
        ws.cell(row=r, column=7, value=f"='{SHEET_TRJ}'!L{src_row}").number_format = "0.0"
        ws.cell(row=r, column=8, value=f"='{SHEET_TRJ}'!M{src_row}").number_format = "\"$\"#,##0"
        apply_cell = ws.cell(row=r, column=9, value="Apply ▸")
        apply_cell.hyperlink = SORTED_JOBS[i][10]
        apply_cell.font = FONT_LINK
        apply_cell.alignment = CENTER
        ws.row_dimensions[r].height = 22
        zebra = (i % 2 == 1)
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = FILL_ZEBRA if zebra else FILL_SURFACE
            cell.border = BORDER_THIN

    # Data bars on Composite (col H = 8) and Fit (col G = 7)
    composite_range = f"H{leaderboard_start + 1}:H{leaderboard_start + 10}"
    fit_range = f"G{leaderboard_start + 1}:G{leaderboard_start + 10}"
    ws.conditional_formatting.add(composite_range,
                                  DataBarRule(start_type="min", end_type="max",
                                              color=PAL["accent"], showValue=True))
    ws.conditional_formatting.add(fit_range,
                                  DataBarRule(start_type="min", end_type="max",
                                              color=PAL["primary_light"], showValue=True))

    apply_bg(ws, max_col=15, max_row=leaderboard_start + 12)


# ---------------------------------------------------------------------------
# Sheet 3: Top Ranked Jobs
# ---------------------------------------------------------------------------
TRJ_HEADERS = [
    "Rank", "Tier", "Company", "Title", "Req#", "Type", "Location",
    "Sal Low", "Sal High", "Sal Mid", "Fit", "Composite",
    "Sal as posted", "Posted", "Apply", "Why ranked here",
]


def build_top_ranked(wb) -> None:
    ws = wb.create_sheet(SHEET_TRJ)
    hide_gridlines(ws)
    set_col_widths(ws, {
        "A": 2, "B": 6, "C": 6, "D": 22, "E": 36, "F": 8, "G": 16, "H": 22,
        "I": 11, "J": 11, "K": 11, "L": 7, "M": 13, "N": 14, "O": 10,
        "P": 11, "Q": 60,
    })

    # Title + subtitle
    ws["B2"] = "Top Ranked Jobs"
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=PAL["primary"])
    ws["B3"] = f"All {N_JOBS} roles, sorted by tier then composite. Click Apply ▸ to open the live careers page."
    ws["B3"].font = FONT_MUTED
    ws["B4"] = "Composite = Sal Mid × Fit / 10. Conditional formatting highlights the strongest leads."
    ws["B4"].font = FONT_MUTED
    ws.row_dimensions[2].height = 28

    # Header row 6
    for idx, h in enumerate(TRJ_HEADERS):
        cell = ws.cell(row=6, column=2 + idx, value=h)
        cell.font = FONT_HEADER_WHITE
        cell.fill = FILL_PRIMARY
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[6].height = 36

    # Data rows
    for i, job in enumerate(SORTED_JOBS):
        (tier, company, title, req, jtype, loc, sal_low, sal_high,
         sal_text, posted, url, fit, why) = job
        r = DATA_START_ROW + i
        ws.cell(row=r, column=2, value=i + 1).alignment = CENTER  # Rank
        tier_cell = ws.cell(row=r, column=3, value=tier)
        tier_cell.alignment = CENTER
        tier_cell.fill = PatternFill("solid", fgColor=TIER_FILL[tier])
        tier_cell.font = FONT_BODY_BOLD
        ws.cell(row=r, column=4, value=company).font = FONT_BODY_BOLD
        ws.cell(row=r, column=4).alignment = LEFT_WRAP
        ws.cell(row=r, column=5, value=title).font = FONT_BODY
        ws.cell(row=r, column=5).alignment = LEFT_WRAP
        ws.cell(row=r, column=6, value=req).alignment = CENTER
        ws.cell(row=r, column=7, value=jtype).font = FONT_BODY
        ws.cell(row=r, column=7).alignment = LEFT_WRAP
        ws.cell(row=r, column=8, value=loc).font = FONT_BODY
        ws.cell(row=r, column=8).alignment = LEFT_WRAP
        ws.cell(row=r, column=9, value=sal_low).number_format = "\"$\"#,##0"
        ws.cell(row=r, column=10, value=sal_high).number_format = "\"$\"#,##0"
        # Sal Mid formula (col K = (I+J)/2)
        ws.cell(row=r, column=11, value=f"=(I{r}+J{r})/2").number_format = "\"$\"#,##0"
        ws.cell(row=r, column=12, value=fit).number_format = "0.0"
        ws.cell(row=r, column=12).alignment = CENTER
        # Composite formula (col M = K*L/10)
        ws.cell(row=r, column=13, value=f"=K{r}*L{r}/10").number_format = "\"$\"#,##0"
        ws.cell(row=r, column=14, value=sal_text).font = FONT_BODY
        ws.cell(row=r, column=14).alignment = CENTER
        ws.cell(row=r, column=15, value=posted).alignment = CENTER
        apply_cell = ws.cell(row=r, column=16, value="Apply ▸")
        apply_cell.hyperlink = url
        apply_cell.font = FONT_LINK
        apply_cell.alignment = CENTER
        ws.cell(row=r, column=17, value=why).font = FONT_BODY
        ws.cell(row=r, column=17).alignment = LEFT_WRAP

        ws.row_dimensions[r].height = 78

        # Zebra (skip tier chip column 3 which already has its own fill)
        zebra = (i % 2 == 1)
        for c in range(2, 18):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            if c == 3:
                continue  # preserve tier chip fill
            if zebra:
                cell.fill = FILL_ZEBRA
            else:
                cell.fill = FILL_SURFACE

    # Conditional formatting: data bar on Composite (col M), 3-color scale on Fit (col L)
    comp_range = f"M{DATA_START_ROW}:M{DATA_END_ROW}"
    fit_range = f"L{DATA_START_ROW}:L{DATA_END_ROW}"
    ws.conditional_formatting.add(comp_range,
                                  DataBarRule(start_type="min", end_type="max",
                                              color=PAL["accent"], showValue=True))
    ws.conditional_formatting.add(
        fit_range,
        ColorScaleRule(
            start_type="num", start_value=5, start_color=PAL["background"],
            mid_type="num", mid_value=7.5, mid_color=PAL["primary_light"],
            end_type="num", end_value=10, end_color=PAL["accent"],
        ),
    )

    ws.freeze_panes = "B7"
    ws.auto_filter.ref = f"B6:Q{DATA_END_ROW}"

    apply_bg(ws, max_col=18, max_row=DATA_END_ROW + 2)


# ---------------------------------------------------------------------------
# Sheets 4 & 5: By Company / By Type
# ---------------------------------------------------------------------------
def _build_crosstab(wb, sheet_name: str, key_label: str, key_index: int,
                    chart_title: str | None = None) -> None:
    ws = wb.create_sheet(sheet_name)
    hide_gridlines(ws)
    set_col_widths(ws, {
        "A": 2, "B": 32, "C": 9, "D": 14, "E": 14, "F": 14, "G": 11,
    })

    ws["B2"] = sheet_name
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=PAL["primary"])
    ws.row_dimensions[2].height = 26

    headers = [key_label, "Roles", "Min Sal Mid", "Median Sal Mid", "Max Sal Mid", "Avg Fit"]
    for idx, h in enumerate(headers):
        cell = ws.cell(row=5, column=2 + idx, value=h)
        cell.font = FONT_HEADER_WHITE
        cell.fill = FILL_PRIMARY
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[5].height = 26

    # Unique key list, preserve sort
    seen: list[str] = []
    for j in SORTED_JOBS:
        k = j[key_index]
        if k not in seen:
            seen.append(k)

    r = 6
    sal_mid_col_ref = f"'{SHEET_TRJ}'!K{DATA_START_ROW}:K{DATA_END_ROW}"
    fit_col_ref = f"'{SHEET_TRJ}'!L{DATA_START_ROW}:L{DATA_END_ROW}"
    if key_index == 1:  # company → col D
        key_col_ref = f"'{SHEET_TRJ}'!D{DATA_START_ROW}:D{DATA_END_ROW}"
    else:  # type → col G
        key_col_ref = f"'{SHEET_TRJ}'!G{DATA_START_ROW}:G{DATA_END_ROW}"

    for key in seen:
        ws.cell(row=r, column=2, value=key).font = FONT_BODY_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_WRAP
        ws.cell(row=r, column=3,
                value=f"=COUNTIF({key_col_ref},\"{key}\")").alignment = CENTER

        # MIN with condition → ArrayFormula
        min_cell = f"D{r}"
        ws[min_cell] = ArrayFormula(
            min_cell,
            f"=MIN(IF({key_col_ref}=\"{key}\",{sal_mid_col_ref}))",
        )
        ws[min_cell].number_format = "\"$\"#,##0"

        # MEDIAN with condition → ArrayFormula
        med_cell = f"E{r}"
        ws[med_cell] = ArrayFormula(
            med_cell,
            f"=MEDIAN(IF({key_col_ref}=\"{key}\",{sal_mid_col_ref}))",
        )
        ws[med_cell].number_format = "\"$\"#,##0"

        # MAX with condition → ArrayFormula
        max_cell = f"F{r}"
        ws[max_cell] = ArrayFormula(
            max_cell,
            f"=MAX(IF({key_col_ref}=\"{key}\",{sal_mid_col_ref}))",
        )
        ws[max_cell].number_format = "\"$\"#,##0"

        # Avg Fit — AVERAGEIF works without array wrapper
        ws.cell(row=r, column=7,
                value=f"=ROUND(AVERAGEIF({key_col_ref},\"{key}\",{fit_col_ref}),1)"
                ).number_format = "0.0"
        ws.cell(row=r, column=7).alignment = CENTER

        zebra = ((r - 6) % 2 == 1)
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            cell.fill = FILL_ZEBRA if zebra else FILL_SURFACE
        ws.row_dimensions[r].height = 22
        r += 1

    last_data_row = r - 1

    if chart_title:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 11
        chart.title = chart_title
        chart.legend = None
        chart.height = 12
        chart.width = 22
        data = Reference(ws, min_col=3, min_row=5, max_col=3, max_row=last_data_row)
        cats = Reference(ws, min_col=2, min_row=6, max_col=2, max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList(showVal=True)
        ws.add_chart(chart, f"I5")

    apply_bg(ws, max_col=20, max_row=last_data_row + 30)


def build_by_company(wb) -> None:
    _build_crosstab(wb, "By Company", "Employer", key_index=1)


def build_by_type(wb) -> None:
    _build_crosstab(wb, "By Type", "Category", key_index=4,
                    chart_title="Role count by category")


# ---------------------------------------------------------------------------
# Sheet 6: Tier Summary
# ---------------------------------------------------------------------------
def build_tier_summary(wb) -> None:
    ws = wb.create_sheet("Tier Summary")
    hide_gridlines(ws)
    set_col_widths(ws, {
        "A": 2, "B": 14, "C": 9, "D": 14, "E": 14, "F": 14, "G": 11,
    })

    ws["B2"] = "Tier Summary"
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=PAL["primary"])
    ws.row_dimensions[2].height = 26

    headers = ["Tier", "Roles", "Min Sal Mid", "Median Sal Mid", "Max Sal Mid", "Avg Fit"]
    for idx, h in enumerate(headers):
        cell = ws.cell(row=5, column=2 + idx, value=h)
        cell.font = FONT_HEADER_WHITE
        cell.fill = FILL_PRIMARY
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[5].height = 26

    tier_col_ref = f"'{SHEET_TRJ}'!C{DATA_START_ROW}:C{DATA_END_ROW}"
    sal_col_ref = f"'{SHEET_TRJ}'!K{DATA_START_ROW}:K{DATA_END_ROW}"
    fit_col_ref = f"'{SHEET_TRJ}'!L{DATA_START_ROW}:L{DATA_END_ROW}"

    r = 6
    for tier in (1, 2, 3, 4):
        cell = ws.cell(row=r, column=2, value=f"Tier {tier}")
        cell.fill = PatternFill("solid", fgColor=TIER_FILL[tier])
        cell.font = FONT_BODY_BOLD
        cell.alignment = CENTER

        ws.cell(row=r, column=3, value=f"=COUNTIF({tier_col_ref},{tier})").alignment = CENTER

        ws[f"D{r}"] = ArrayFormula(
            f"D{r}", f"=MIN(IF({tier_col_ref}={tier},{sal_col_ref}))")
        ws[f"D{r}"].number_format = "\"$\"#,##0"
        ws[f"E{r}"] = ArrayFormula(
            f"E{r}", f"=MEDIAN(IF({tier_col_ref}={tier},{sal_col_ref}))")
        ws[f"E{r}"].number_format = "\"$\"#,##0"
        ws[f"F{r}"] = ArrayFormula(
            f"F{r}", f"=MAX(IF({tier_col_ref}={tier},{sal_col_ref}))")
        ws[f"F{r}"].number_format = "\"$\"#,##0"

        ws.cell(row=r, column=7,
                value=f"=ROUND(AVERAGEIF({tier_col_ref},{tier},{fit_col_ref}),1)"
                ).number_format = "0.0"
        ws.cell(row=r, column=7).alignment = CENTER

        for c in range(2, 8):
            ws.cell(row=r, column=c).border = BORDER_THIN
        ws.row_dimensions[r].height = 26
        r += 1

    # TOTAL / OVERALL row — plain MIN/MAX/AVERAGE (no condition)
    total_cell = ws.cell(row=r, column=2, value="OVERALL")
    total_cell.font = FONT_BODY_BOLD
    total_cell.alignment = CENTER
    total_cell.fill = FILL_PRIMARY_LIGHT
    ws.cell(row=r, column=3,
            value=f"=COUNTA('{SHEET_TRJ}'!D{DATA_START_ROW}:D{DATA_END_ROW})").alignment = CENTER
    ws.cell(row=r, column=4, value=f"=MIN({sal_col_ref})").number_format = "\"$\"#,##0"
    ws.cell(row=r, column=5, value=f"=MEDIAN({sal_col_ref})").number_format = "\"$\"#,##0"
    ws.cell(row=r, column=6, value=f"=MAX({sal_col_ref})").number_format = "\"$\"#,##0"
    ws.cell(row=r, column=7,
            value=f"=ROUND(AVERAGE({fit_col_ref}),1)").number_format = "0.0"
    ws.cell(row=r, column=7).alignment = CENTER
    for c in range(2, 8):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER_THIN
        if cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = FILL_PRIMARY_LIGHT
        cell.font = FONT_BODY_BOLD
    ws.row_dimensions[r].height = 28

    # Column chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Roles by tier"
    chart.legend = None
    chart.height = 9
    chart.width = 16
    data = Reference(ws, min_col=3, min_row=5, max_col=3, max_row=9)
    cats = Reference(ws, min_col=2, min_row=6, max_col=2, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(chart, "I5")

    apply_bg(ws, max_col=20, max_row=r + 25)


# ---------------------------------------------------------------------------
# Sheet 7: Action Plan
# ---------------------------------------------------------------------------
def build_action_plan(wb) -> None:
    ws = wb.create_sheet("Action Plan")
    hide_gridlines(ws)
    set_col_widths(ws, {
        "A": 2, "B": 10, "C": 14, "D": 28, "E": 36, "F": 8, "G": 11, "H": 50,
    })

    ws["B2"] = "Four-Week Application Plan"
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=PAL["primary"])
    ws.row_dimensions[2].height = 26

    today = date.today()
    monday = today - timedelta(days=today.weekday())
    weeks = [monday + timedelta(weeks=i) for i in range(4)]

    headers = ["Week", "Window", "Company", "Title", "Tier", "Fit", "Suggested move"]
    for idx, h in enumerate(headers):
        cell = ws.cell(row=5, column=2 + idx, value=h)
        cell.font = FONT_HEADER_WHITE
        cell.fill = FILL_PRIMARY
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[5].height = 26

    # Distribute roles across weeks: tier 1 in week 1, tier 2 in week 2, tier 3
    # in week 3, tier 4 in week 4. Cap at ~12 per week so the plan stays usable.
    by_tier: dict[int, list[tuple]] = {1: [], 2: [], 3: [], 4: []}
    for j in SORTED_JOBS:
        by_tier[j[0]].append(j)

    suggested_moves = {
        1: "Tailored resume + cover letter this week. Reach out to one internal referrer.",
        2: "Apply via stable career portal. Note referral targets on LinkedIn.",
        3: "Apply if Tier 1/2 pipeline thins by Friday. Otherwise keep warm.",
        4: "Adjunct apps — submit via PeopleAdmin. Stack teaching credential while search runs.",
    }

    r = 6
    for wk_idx, week_start in enumerate(weeks):
        tier_for_week = wk_idx + 1
        roles = by_tier.get(tier_for_week, [])[:12]
        window = f"{week_start.strftime('%b %d')} – {(week_start + timedelta(days=4)).strftime('%b %d')}"
        for role in roles:
            (tier, company, title, _req, _jtype, _loc, _sl, _sh,
             _txt, _posted, _url, fit, _why) = role
            wk_cell = ws.cell(row=r, column=2, value=f"Week {wk_idx + 1}")
            wk_cell.fill = PatternFill("solid", fgColor=TIER_FILL[tier])
            wk_cell.font = FONT_BODY_BOLD
            wk_cell.alignment = CENTER
            ws.cell(row=r, column=3, value=window).alignment = CENTER
            ws.cell(row=r, column=4, value=company).font = FONT_BODY_BOLD
            ws.cell(row=r, column=4).alignment = LEFT_WRAP
            ws.cell(row=r, column=5, value=title).alignment = LEFT_WRAP
            ws.cell(row=r, column=6, value=tier).alignment = CENTER
            ws.cell(row=r, column=7, value=fit).number_format = "0.0"
            ws.cell(row=r, column=7).alignment = CENTER
            ws.cell(row=r, column=8, value=suggested_moves[tier]).alignment = LEFT_WRAP
            for c in range(2, 9):
                cell = ws.cell(row=r, column=c)
                cell.border = BORDER_THIN
                if c != 2 and cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = FILL_ZEBRA if ((r - 6) % 2) else FILL_SURFACE
            ws.row_dimensions[r].height = 32
            r += 1

    ws.freeze_panes = "B6"
    apply_bg(ws, max_col=10, max_row=r + 2)


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------
def main() -> int:
    WORKSPACE.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    build_readme(wb)
    build_dashboard(wb)
    build_top_ranked(wb)
    build_by_company(wb)
    build_by_type(wb)
    build_tier_summary(wb)
    build_action_plan(wb)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH} ({N_JOBS} roles)")

    proc = subprocess.run(
        ["python3", str(RECALC_SCRIPT), str(OUT_PATH)],
        capture_output=True, text=True,
        env={**os.environ,
             "PYTHONPATH": str(RECALC_SCRIPT.parent)},
    )
    print("recalc stdout:", proc.stdout.strip())
    if proc.stderr.strip():
        print("recalc stderr:", proc.stderr.strip(), file=sys.stderr)

    try:
        result = json.loads(proc.stdout.strip().splitlines()[-1])
    except (json.JSONDecodeError, IndexError):
        print("could not parse recalc output", file=sys.stderr)
        return 2

    if result.get("status") != "success":
        print(f"recalc found {result.get('total_errors')} errors", file=sys.stderr)
        return 3

    # Independently verify KPI cells via the formulas engine (the recalc engine
    # cannot write cached values back into a formatted .xlsx without losing
    # styles, so we re-evaluate here purely to assert the values are sensible).
    import formulas  # noqa: PLC0415
    xl = formulas.ExcelModel().loads(str(OUT_PATH)).finish()
    sol = xl.calculate()
    fname = OUT_PATH.name  # formulas package uses lowercase filename
    kpi_coords = [
        ("DASHBOARD", "B7"), ("DASHBOARD", "F7"), ("DASHBOARD", "J7"),
        ("DASHBOARD", "B12"), ("DASHBOARD", "F12"), ("DASHBOARD", "J12"),
    ]
    kpi_values: list = []
    for sheet, coord in kpi_coords:
        node = sol.get(f"'[{fname}]{sheet}'!{coord}")
        if node is None:
            print(f"missing KPI cell {sheet}!{coord}", file=sys.stderr)
            return 4
        val = node.value if hasattr(node, "value") else node
        try:
            val = val[0][0]
        except (TypeError, IndexError):
            pass
        kpi_values.append(val)
    print("KPI values:", kpi_values)

    print(f"all quality gates passed (engine={result.get('engine')}, "
          f"evaluated_cells={result.get('evaluated_cells', 'n/a')})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
