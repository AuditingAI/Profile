"""Build the Recruiting Launch Checklist xlsx referenced in the README
but not yet on the repo. Matches the styling of the existing Pilot
Feedback Form xlsx. Five gate tabs.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "YMalik_Recruiting_Launch_Checklist_READY_2026-06-01.xlsx"
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="0C4E54")
SUB_FONT = Font(name="Calibri", size=10, italic=True, color="6F6F6F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
HEADER_FILL = PatternFill("solid", fgColor="0C4E54")
ZEBRA = PatternFill("solid", fgColor="F6F4EE")
THIN = Side(style="thin", color="CFCBC0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def setup_sheet(ws, title, subtitle, headers, widths, body_rows):
    ws["A1"] = title; ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle; ws["A2"].font = SUB_FONT
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 30
    for r_idx, row in enumerate(body_rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT; cell.alignment = WRAP; cell.border = BOX
            if r_idx % 2 == 0:
                cell.fill = ZEBRA
        ws.row_dimensions[r_idx].height = 36

wb = Workbook()
ws = wb.active; ws.title = "1. Pre-Pilot Gate"
setup_sheet(ws, "Pre-Pilot Readiness Gate",
    "All items must be GREEN before the informed pilot link is shared with any participant.",
    ["#", "Gate Item", "Owner", "Evidence Required", "Status", "Date Cleared", "Notes"],
    [4, 44, 12, 38, 12, 14, 30],
    [
        (1, "Qualtrics build matches IRB-approved instrument (IRB-25-0462); deltas filed as non-substantive change or amendment.", "Yasir", "Side-by-side check vs. IRB-approved PDF; advisor sign-off note.", "Pending", "", ""),
        (2, "Consent block in Qualtrics presents the full informational letter and a forced-choice consent question.", "Yasir", "Screenshot of the consent page from a clean preview link.", "Pending", "", ""),
        (3, "Eligibility screener filters non-eligible respondents into End-of-Survey before the construct items load.", "Yasir", "Qualtrics survey flow screenshot showing branching.", "Pending", "", ""),
        (4, "Two attention checks embedded (per Appendix A.4 and A.8) and configured to mark failures, not auto-terminate.", "Yasir", "Qualtrics logic screenshot; results variable defined.", "Pending", "", ""),
        (5, "Reverse-coded items render correctly and the data dictionary captures direction.", "Yasir", "Item-by-item export from Qualtrics; reverse-coding key.", "Pending", "", ""),
        (6, "Progress bar enabled; estimated completion time displayed at start.", "Yasir", "Preview screenshot from start of survey.", "Pending", "", ""),
        (7, "Mobile-responsive layout tested on phone and tablet.", "Yasir", "Two screenshots: phone preview, tablet preview.", "Pending", "", ""),
        (8, "No personally identifying or client-identifying fields collected (consistent with IRB scope).", "Yasir", "Field-by-field audit against IRB approval.", "Pending", "", ""),
        (9, "Advisor face-validity sign-off on full Qualtrics build.", "Dr. Rey", "Email or meeting note from Dr. Rey confirming sign-off.", "Pending", "", ""),
        (10, "Pilot Participant Log (Pilot_Feedback xlsx, Tab 1) opened and ready for entries.", "Yasir", "File saved at 03_Recruitment_and_Pilot/ with first row populated.", "Pending", "", ""),
    ])

ws = wb.create_sheet("2. Pilot Close-Out Gate")
setup_sheet(ws, "Pilot Close-Out Gate",
    "All items must be GREEN before the recruiting platform listing goes live.",
    ["#", "Gate Item", "Threshold / Decision Rule", "Status", "Date Cleared", "Notes"],
    [4, 44, 38, 12, 14, 30],
    [
        (1, "Pilot sample size achieved.", "n = 6-10 eligible auditors completed both the survey and the feedback form.", "Pending", "", ""),
        (2, "Median completion time within target.", "15-20 minutes; flag if outside band.", "Pending", "", ""),
        (3, "No respondent reported items that were unclear, leading, or domain-inappropriate.", "PF1-PF4 mean >= 4.0 on the 1-5 scale across pilot respondents.", "Pending", "", ""),
        (4, "No structural item changes required.", "Severity column on Revision Log shows no entries flagged High and no entries flagged 'Advisor/IRB Review Needed = Yes'.", "Pending", "", ""),
        (5, "Attention-check failure rate within tolerance.", "<= 10% of pilot respondents failed either attention check.", "Pending", "", ""),
        (6, "Pilot Decision Summary tab fully populated and all criteria status = Met.", "All criteria on the Pilot Decision Summary tab = Met.", "Pending", "", ""),
        (7, "Advisor sign-off to proceed to recruiting platform.", "Dr. Rey confirms via email or meeting note.", "Pending", "", ""),
        (8, "Any IRB amendment triggered by pilot revisions has been filed and acknowledged.", "Topaz confirmation number on file; if no amendment required, note that here.", "Pending", "", ""),
    ])

ws = wb.create_sheet("3. CloudResearch Gate")
setup_sheet(ws, "CloudResearch Launch Gate (Primary Recruiting Route)",
    "Gate Items 1-8 must be GREEN before the CloudResearch study listing is set to Live. Gate Items 9-12 govern the soft launch.",
    ["#", "Gate Item", "Owner", "Evidence Required", "Status", "Date Cleared", "Notes"],
    [4, 44, 12, 38, 12, 14, 30],
    [
        (1, "Study listing copy reviewed against the CloudResearch Launch Draft (READY 2026-06-01).", "Yasir", "Diff vs. approved draft.", "Pending", "", ""),
        (2, "Compensation amount confirmed with advisor and within CloudResearch fair-pay band.", "Yasir + Dr. Rey", "Advisor confirmation; cross-check vs. CR fair-pay calculator.", "Pending", "", ""),
        (3, "Quality filters configured: approval rating, prior-participation exclusion, geo, audit-profession screener.", "Yasir", "Screenshot of CR filter panel.", "Pending", "", ""),
        (4, "Do-not-add boundaries verbalized in the listing (no AI/LLM constructs, no fabricated identity).", "Yasir", "Listing copy review.", "Pending", "", ""),
        (5, "Time estimate in the listing matches pilot median completion time.", "Yasir", "Pilot data + listing copy.", "Pending", "", ""),
        (6, "Survey link routes from CR to Qualtrics with the CR participant ID embedded.", "Yasir", "Test link round-trip: CR -> Qualtrics -> CR completion code.", "Pending", "", ""),
        (7, "Completion code returned to CR for crediting; failure path returns a distinguishable code.", "Yasir", "Screenshot of CR validation panel; two test runs.", "Pending", "", ""),
        (8, "Data export path validated: SPSS / CSV / SAV download from Qualtrics works for at least one test response.", "Yasir", "Downloaded test file saved locally.", "Pending", "", ""),
        (9, "Soft launch n = 10-15 only; pause for 24 hours to inspect data quality.", "Yasir", "Date/time of soft-launch open and close.", "Pending", "", ""),
        (10, "Soft-launch quality review: attention-check pass rate, straight-lining check, completion-time outliers.", "Yasir", "One-page review note saved to correspondence/.", "Pending", "", ""),
        (11, "Decision to proceed to full launch (n = 60-80 target) or pause for fixes.", "Yasir + Dr. Rey", "Advisor acknowledgement.", "Pending", "", ""),
        (12, "Full-launch ongoing data quality monitoring cadence agreed (every 20 responses).", "Yasir", "Cadence noted in this sheet.", "Pending", "", ""),
    ])

ws = wb.create_sheet("4. MTurk Backup Gate")
setup_sheet(ws, "MTurk Backup Gate (Conditional — only if CloudResearch yield insufficient)",
    "Do not open MTurk concurrently with CloudResearch. Open only if CR yields below target after the full-launch window.",
    ["#", "Gate Item", "Owner", "Evidence Required", "Status", "Date Cleared", "Notes"],
    [4, 44, 12, 38, 12, 14, 30],
    [
        (1, "CloudResearch yield confirmed below target after defined full-launch window.", "Yasir", "Yield report from CR; advisor confirmation that backup is needed.", "Pending", "", ""),
        (2, "MTurk HIT copy reviewed against the MTurk Backup Launch Draft (READY 2026-06-01).", "Yasir", "Diff vs. approved draft.", "Pending", "", ""),
        (3, "Qualifications set: HIT approval rate, master qualification, audit screener.", "Yasir", "Screenshot of HIT qualifications panel.", "Pending", "", ""),
        (4, "Duplicate-participant exclusion: CR completion IDs added to MTurk exclusion list.", "Yasir", "Exclusion list confirmed.", "Pending", "", ""),
        (5, "Per-assignment reward consistent with CR rate.", "Yasir", "Cross-check.", "Pending", "", ""),
        (6, "Same survey-link round trip and completion-code validation as in the CR gate.", "Yasir", "Two test runs.", "Pending", "", ""),
        (7, "Soft launch n = 5-10 on MTurk first.", "Yasir", "Date/time of soft-launch open and close.", "Pending", "", ""),
        (8, "Decision to scale or pause MTurk arm; advisor sign-off.", "Yasir + Dr. Rey", "Advisor acknowledgement.", "Pending", "", ""),
    ])

ws = wb.create_sheet("5. Stop Conditions")
setup_sheet(ws, "Stop / Pause Conditions During Live Recruitment",
    "If any condition triggers, pause data collection on the affected route within 24 hours and document the pause in the Revision Log.",
    ["#", "Condition", "Detection Method", "Pause Action", "Owner", "Triggered? (Y/N)", "Notes"],
    [4, 38, 32, 32, 12, 14, 26],
    [
        (1, "Attention-check failure rate exceeds 15% across the last 20 responses.", "Rolling check from Qualtrics export.", "Pause CR / MTurk; review item wording.", "Yasir", "N", ""),
        (2, "Straight-lining detected in > 10% of responses.", "Variance < 0.5 across construct items.", "Pause; consider adding a second behavioral check.", "Yasir", "N", ""),
        (3, "Median completion time drops below 8 minutes or exceeds 35 minutes.", "Qualtrics duration export.", "Pause; review the listing time estimate.", "Yasir", "N", ""),
        (4, "Eligibility-screener pass rate falls below 30%.", "Eligible / Total in Qualtrics.", "Adjust screener wording; refile if material.", "Yasir", "N", ""),
        (5, "Any respondent reports a privacy or sensitive-data concern.", "Feedback channel monitored daily.", "Pause immediately; report to FIU IRB if applicable.", "Yasir", "N", ""),
        (6, "Platform-level concern flagged by CR or MTurk worker community.", "Daily check of platform-specific worker forums.", "Pause the affected route; consult advisor.", "Yasir", "N", ""),
    ])

for s in wb.worksheets:
    s.freeze_panes = "A5"
    s.sheet_view.showGridLines = False

wb.save(OUT)
print("wrote", OUT)
